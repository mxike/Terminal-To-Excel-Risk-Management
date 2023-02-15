import os, time
from os import system
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()
wb = load_workbook(os.getenv('FILE_PATH'))

class Tab:
  def showSheets(self):
    system('cls')
    print("Active worksheet: ", wb.active)
    print("All worksheets: ", wb.sheetnames)
    print()
    self.menu()

  def addSheet(self):
    system('cls')
    option = int(input("(1) Manual - (2) Automatic: "))
    if(option == 1 or option == "manual"):
      sheet_name = input("Choose a sheet name: ")
      wb.create_sheet(sheet_name.capitalize())
      wb.save(os.getenv('FILE_PATH'))
      time.sleep(.2)
      print("Succesfully added: {}".format(sheet_name))
      print()

    elif(option == 2 or option == "automatic"):
      months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "Oktober", "November", "December"]
      added = []
      for month in months:
        if month not in wb.sheetnames:
          added.append(month)
          wb.create_sheet(month)
      wb.save(os.getenv('FILE_PATH'))
      time.sleep(.2)
      print("Succesfully added: {}".format(added))
      print()
    else:
      system('cls')
      self.menu()

    self.menu()

  def change_Active_Sheet(self):
    system('cls')
    print("Active worksheet: ", wb.active)
    print("All worksheets: ", wb.sheetnames)

    input_change_active = input("Choose new active sheet: ").capitalize()
    if input_change_active in wb.sheetnames:
      wb.active = wb[input_change_active]
      print("New active sheet: ", wb.active)
      wb.save(os.getenv("FILE_PATH"))
    else:
      print("Sheet:", input_change_active, "- Doesn't exist")

    print()
    self.menu()

  def removeSheet(self):
    system('cls')
    print("All worksheets: ", wb.sheetnames)
    option = int(input("(1) Remove one sheet (2) Remove all sheets except active sheet: "))

    if option == 1:
      delete_sheet_input = input("Select which sheet you want to delete: ").capitalize()
      if delete_sheet_input not in wb.sheetnames:
        wb.remove_sheet(wb.get_sheet_by_name(delete_sheet_input))
        wb.save(os.getenv("FILE_PATH"))          
      else:
        print("Sheet name doesn't exist")

    elif option == 2:
      print("This option will remove all sheets except the active one: ", wb.active)
      sure_input = input("are you sure? Y/N: ").upper()
      if sure_input == "Y" or sure_input == "YES":
        active_sheet = str(wb.active)
        for sheet in wb.sheetnames:
          if active_sheet[12:-2] != sheet:
            wb.remove_sheet(wb.get_sheet_by_name(sheet))
            wb.save(os.getenv("FILE_PATH"))
        time.sleep(.2)
        print("Succesfully removed all sheets")
        print("Remaining sheets: ", wb.sheetnames)

    else:
      print("Option doens't exist")

    print()
    self.menu()

  def back(self):
    from main import main_menu
    main_menu()

  def menu(self):
    from main import create_menu, menu_title
    menu_title("Tab Menu", "bright_blue")
    tab_menu = [self.showSheets, self.addSheet, self.change_Active_Sheet, self.removeSheet, self.back]

    create_menu(tab_menu)
    